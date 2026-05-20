import pandas as pd
import numpy as np
import time
import os
import csv
from datetime import datetime
from cg_behavior import column_generation_behavior
from worker_groups import create_homogeneous_group
from nonlinear_transitions import evaluate_schedule_nl, h_func, r_func
from Utils.gcutil import generate_dict_from_excel, analyze_and_plot_blocks
from Utils.metrics import evaluate_inequality
from Utils.aggundercover import create_dict_from_list, dict_reducer

def get_matrices(epsilon):
    """Defines the three delta matrices."""
    base = np.zeros((4, 4))
    base[1,1]=1.0; base[1,2]=1.2; base[1,3]=1.5
    base[2,1]=2.5; base[2,2]=1.0; base[2,3]=1.2
    base[3,1]=1.2; base[3,2]=1.5; base[3,3]=1.0
    
    severe = np.zeros((4, 4))
    severe[1,1]=1.0; severe[1,2]=2.0; severe[1,3]=2.5
    severe[2,1]=5.0; severe[2,2]=1.0; severe[2,3]=2.0
    severe[3,1]=2.0; severe[3,2]=2.5; severe[3,3]=1.0
    
    close = np.zeros((4, 4))
    close[1,1]=1.0; close[1,2]=1.1; close[1,3]=1.2
    close[2,1]=1.5; close[2,2]=1.0; close[2,3]=1.1
    close[3,1]=1.1; close[3,2]=1.2; close[3,3]=1.0
    
    # Apply epsilon scaling
    for m in [base, severe, close]:
        m *= epsilon
        
    return {
        'base': base,
        'severe': severe,
        'close': close
    }

def check_and_backup_file(file_path, expected_headers, is_excel=False):
    """Check if file exists and has matching headers. If not, backs it up."""
    if not os.path.isfile(file_path):
        return False
        
    backup_needed = False
    try:
        if is_excel:
            from openpyxl import load_workbook
            wb = load_workbook(file_path, read_only=True)
            ws = wb.active
            first_row = next(ws.iter_rows(max_row=1, values_only=True))
            existing_headers = [str(cell).strip() for cell in first_row if cell is not None]
            if len(existing_headers) != len(expected_headers) or any(h1 != h2 for h1, h2 in zip(existing_headers, expected_headers)):
                backup_needed = True
        else:
            with open(file_path, mode='r') as f:
                reader = csv.reader(f)
                first_row = next(reader)
                existing_headers = [cell.strip() for cell in first_row]
                if len(existing_headers) != len(expected_headers) or any(h1 != h2 for h1, h2 in zip(existing_headers, expected_headers)):
                    backup_needed = True
    except Exception as e:
        print(f"Error checking header for {file_path}: {e}. Backing up.")
        backup_needed = True
        
    if backup_needed:
        backup_path = file_path.rsplit('.', 1)[0] + f"_backup_{datetime.now().strftime('%d_%m_%Y_%H-%M')}.xlsx" if is_excel else file_path.rsplit('.', 1)[0] + f"_backup_{datetime.now().strftime('%d_%m_%Y_%H-%M')}.csv"
        try:
            os.rename(file_path, backup_path)
            print(f"Existing file {file_path} had different headers. Backed up to {backup_path}")
        except Exception as e:
            print(f"Could not rename file {file_path}: {e}")
        return False
    return True

def evaluate_naive_expost_nl(ls_x_naive, T, K, nl_spec, demand, I):
    """Ex-post evaluation of naive schedule under nonlinear fatigue model."""
    n_workers = len(I)
    n_days = len(T)
    n_shifts = len(K)
    sublist_length = n_days * n_shifts
    
    ls_x_naive_expost = [1.0 if x > 0.5 else 0.0 for x in ls_x_naive]
    
    ls_perf_naive_expost = []
    ls_p_naive_expost = []
    ls_sc_naive_expost = []
    ls_rec_naive_expost = []
    
    real_supply_per_shift = [0.0] * (n_days * n_shifts)
    
    for idx in range(n_workers):
        worker_x = ls_x_naive_expost[idx * sublist_length : (idx + 1) * sublist_length]
        x_dict = {}
        for d in range(n_days):
            for s in range(n_shifts):
                if worker_x[d * n_shifts + s] > 0.5:
                    x_dict[(T[d], K[s])] = 1.0
                    
        p_hist, sc_hist, r_hist, _ = evaluate_schedule_nl(x_dict, T, K, nl_spec)
        
        for d in range(n_days):
            day = T[d]
            p = p_hist[day]
            sc = sc_hist[day]
            r = r_hist[day]
            
            ls_p_naive_expost.append(p)
            ls_sc_naive_expost.append(sc)
            ls_rec_naive_expost.append(r)
            
            for s in range(n_shifts):
                if worker_x[d * n_shifts + s] > 0.5:
                    ls_perf_naive_expost.append(p)
                    real_supply_per_shift[d * n_shifts + s] += p
                else:
                    ls_perf_naive_expost.append(0.0)
                    
    demand_vals = [demand.get((t, k), 0) for t in T for k in K]
    
    # undercover
    shift_undercover_naive_list = [max(0.0, demand_vals[i] - real_supply_per_shift[i]) for i in range(len(demand_vals))]
    undercover_naive = sum(shift_undercover_naive_list)
    
    # understaffing (nominal missing)
    nominal_supply = [0.0] * (n_days * n_shifts)
    for i in range(len(ls_x_naive_expost)):
        if ls_x_naive_expost[i] > 0.5:
            nominal_supply[i % (n_days * n_shifts)] += 1.0
    understaffing_naive = sum(max(0, demand_vals[i] - nominal_supply[i]) for i in range(len(demand_vals)))
    
    perf_naive = undercover_naive - understaffing_naive
    consistency_naive = sum(ls_sc_naive_expost)
    
    scale = 1.0
    undercover_norm_naive = undercover_naive / (n_workers * scale)
    understaffing_norm_naive = understaffing_naive / (n_workers * scale)
    perf_norm_naive = perf_naive / (n_workers * scale)
    cons_norm_naive = consistency_naive / (n_workers * scale)
    
    shift_undercover_naive = [0.0 if abs(round(x, 3)) == 0 else round(x, 2) for x in shift_undercover_naive_list]
    
    shift_uc_naive_dict = create_dict_from_list(shift_undercover_naive, n_days, n_shifts)
    daily_undercover_naive = dict_reducer(shift_uc_naive_dict)
    
    # Inequality
    L_perf = [x * (1.0 - p) for x, p in zip(ls_x_naive_expost, ls_perf_naive_expost)]
    results_ineq_sc_naive, spread_sc_naive, load_share_sc_naive, gini_sc_naive = evaluate_inequality(ls_sc_naive_expost, n_days, n_workers)
    results_ineq_perf_naive, spread_perf_naive, load_share_perf_naive, gini_perf_naive = evaluate_inequality(
        [sum(L_perf[i:i + 3]) for i in range(0, len(L_perf), 3)], n_days, n_workers
    )
    
    # shift blocks
    shift_blocks_naive = analyze_and_plot_blocks(ls_x_naive_expost, n_workers, n_days, n_shifts)
    
    return {
        'undercover_naive': undercover_naive,
        'undercover_norm_naive': undercover_norm_naive,
        'cons_naive': consistency_naive,
        'cons_norm_naive': cons_norm_naive,
        'perf_naive': perf_naive,
        'perf_norm_naive': perf_norm_naive,
        'understaffing_naive': understaffing_naive,
        'understaffing_norm_naive': understaffing_norm_naive,
        'shift_undercover_naive': shift_undercover_naive,
        'perf_list_naive': ls_perf_naive_expost,
        'cons_list_naive': ls_sc_naive_expost,
        'p_list_naive': ls_p_naive_expost,
        'x_list_naive': ls_x_naive_expost,
        'r_list_naive': ls_rec_naive_expost,
        'daily_undercover_naive': daily_undercover_naive,
        'results_ineq_sc_naive': results_ineq_sc_naive,
        'spread_sc_naive': spread_sc_naive,
        'load_share_sc_naive': load_share_sc_naive,
        'gini_sc_naive': gini_sc_naive,
        'results_ineq_perf_naive': results_ineq_perf_naive,
        'spread_perf_naive': spread_perf_naive,
        'load_share_perf_naive': load_share_perf_naive,
        'gini_perf_naive': gini_perf_naive,
        'shift_blocks_naive': shift_blocks_naive
    }

def run_experiment():
    I_sizes = [100]
    scarcities = [('Medium', 1.0)]
    seeds = range(1,2)
    
    # Parameter Sets
    epsilon = 0.06
    matrices_dict = get_matrices(epsilon)
    chis = [1, 3, 5, 7]
    alphas = [0.02, 0.04, 0.08]
    
    steps_convex = [1.2, 1.5, 2.0]
    steps_concave = [0.5, 0.7, 0.9]
    
    types = {
        'conv': (steps_convex, steps_convex),
        'conc': (steps_concave, steps_concave),
        'convconc': (steps_concave, steps_convex),
        'concconv': (steps_convex, steps_concave)
    }
    
    from openpyxl import Workbook, load_workbook

    range_suffix = f"range_{seeds[0]}_{seeds[-1]}" if len(seeds) > 0 else "empty"
    csv_file = f"sensitivity_results_100_med_{range_suffix}.csv"
    excel_file = f"sensitivity_results_100_med_{range_suffix}.xlsx"
    
    headers = [
        # Sensitivity study parameters
        'seed', 'matrix', 'chi', 'alpha', 'type', 'gamma_r', 'gamma_c', 
        'bap_obj', 'npp_nom', 'npp_expost', 'improvement_pct', 'time',
        # loop_cg columns
        'I', 'T', 'K', 'pattern', 'scenario', 'prob', 'epsilon',
        'gap', 'lagrange', 'objval', 'lbound', 'iteration', 'time_sp', 'time_rmp', 'time_ip',
        'undercover_behavior', 'undercover_norm_behavior', 'cons_behavior', 'cons_norm_behavior', 'perf_behavior', 'perf_norm_behavior',
        'understaffing_behavior', 'understaffing_norm_behavior',
        'undercover_naive', 'undercover_norm_naive', 'cons_naive', 'cons_norm_naive', 'perf_naive', 'perf_norm_naive',
        'understaffing_naive', 'understaffing_norm_naive',
        'shift_undercover_naive', 'shift_undercover_behavior',
        'perf_list_behavior', 'perf_list_naive', 'cons_list_behavior', 'cons_list_naive',
        'p_list_behavior', 'p_list_naive', 'x_list_behavior', 'x_list_naive', 'r_list_behavior', 'r_list_naive',
        'daily_undercover_behavior', 'daily_undercover_naive',
        'results_ineq_sc_behavior', 'results_ineq_sc_naive', 'spread_sc_behavior', 'spread_sc_naive',
        'load_share_sc_behavior', 'load_share_sc_naive', 'gini_sc_behavior', 'gini_sc_naive',
        'results_ineq_perf_behavior', 'results_ineq_perf_naive', 'spread_perf_behavior', 'spread_perf_naive',
        'load_share_perf_behavior', 'load_share_perf_naive', 'gini_perf_behavior', 'gini_perf_naive',
        'shift_blocks_behavior', 'shift_blocks_naive'
    ]

    file_exists = check_and_backup_file(csv_file, headers, is_excel=False)
    excel_exists = check_and_backup_file(excel_file, headers, is_excel=True)
    
    if excel_exists:
        wb = load_workbook(excel_file)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        wb.save(excel_file)
        
    with open(csv_file, mode='a', newline='') as f:
        writer = csv.writer(f)
        if not file_exists:
            writer.writerow(headers)
        
        total_configs = len(matrices_dict) * len(chis) * len(alphas) * 4 * 9
        total_runs = total_configs * len(seeds)
        count = 0
        
        for seed in seeds:
            # LOAD DEMAND FROM EXCEL (AS IN OTHER SCRIPTS)
            demand = generate_dict_from_excel('data/demand_data.xlsx', 100, 'Medium', scenario=seed)
            
            if not demand:
                print(f"Skipping Seed {seed} - No demand data found.")
                continue

            T_days = 28
            T = list(range(1, T_days + 1))
            K = [1, 2, 3]
            I = list(range(1, 101))
            
            df_placeholder = pd.DataFrame({
                'I': I + [np.nan] * (max(len(I), len(T), len(K)) - len(I)),
                'T': T + [np.nan] * (max(len(I), len(T), len(K)) - len(T)),
                'K': K + [np.nan] * (max(len(I), len(T), len(K)) - len(K))
            })

            for m_name, delta in matrices_dict.items():
                for chi in chis:
                    for alpha in alphas:
                        for t_name, (r_steps, c_steps) in types.items():
                            for gr in r_steps:
                                for gc in c_steps:
                                    count += 1
                                    if count % 10 == 0 or count == 1:
                                        print(f"Run {count}/{total_runs} (Seed {seed}): {m_name}, chi={chi}, {t_name}")
                                    
                                    nl_spec = {
                                        'epsilon': epsilon,
                                        'chi': chi,
                                        'gamma_R': gr,
                                        'gamma_C': gc,
                                        'alpha_R': alpha,
                                        'delta': delta,
                                        'e_max': 0.5
                                    }
                                    
                                    start_t = time.time()
                                    
                                    # 1. BAP
                                    bap_res = column_generation_behavior(
                                        df_placeholder, demand, epsilon, 4, 6, 10, 100, 0, chi, 0.001, 10, I, T, K, 1.0,
                                        sp_solver="labeling", nl_spec=nl_spec
                                    )
                                    bap_obj = bap_res[8]
                                    
                                    # 2. NPP
                                    npp_res = column_generation_behavior(
                                        df_placeholder, demand, 0.0, 4, 6, 10, 100, 0, chi, 0.001, 10, I, T, K, 1.0,
                                        sp_solver="labeling", nl_spec=None
                                    )
                                    npp_nom = npp_res[8]
                                    
                                    # 3. NPP Ex-Post
                                    ls_x_naive = npp_res[19]
                                    naive_expost_metrics = evaluate_naive_expost_nl(ls_x_naive, T, K, nl_spec, demand, I)
                                    npp_expost = naive_expost_metrics['undercover_naive']
                                    
                                    imp = ((npp_expost - bap_obj) / npp_expost * 100) if npp_expost > 0 else 0
                                    elapsed = time.time() - start_t
                                    
                                    # Construct daily undercover dict reducer for behavior
                                    shift_uc_behavior_dict = create_dict_from_list(bap_res[21], len(T), len(K))
                                    daily_undercover_behavior = dict_reducer(shift_uc_behavior_dict)
                                    
                                    row_dict = {
                                        # Sensitivity study parameters
                                        'seed': seed,
                                        'matrix': m_name,
                                        'chi': chi,
                                        'alpha': alpha,
                                        'type': t_name,
                                        'gamma_r': gr,
                                        'gamma_c': gc,
                                        'bap_obj': bap_obj,
                                        'npp_nom': npp_nom,
                                        'npp_expost': npp_expost,
                                        'improvement_pct': imp,
                                        'time': elapsed,
                                        # loop_cg columns
                                        'I': len(I),
                                        'T': len(T),
                                        'K': len(K),
                                        'pattern': 'Medium',
                                        'scenario': seed,
                                        'prob': 1.0,
                                        'epsilon': epsilon,
                                        'gap': round(bap_res[12], 3),
                                        'lagrange': round(bap_res[11], 3),
                                        'objval': round(bap_res[8], 3),
                                        'lbound': round(bap_res[9], 3),
                                        'iteration': bap_res[10],
                                        'time_sp': round(bap_res[13], 3),
                                        'time_rmp': round(bap_res[14], 3),
                                        'time_ip': round(bap_res[15], 3),
                                        'undercover_behavior': bap_res[0],
                                        'undercover_norm_behavior': bap_res[5],
                                        'cons_behavior': bap_res[3],
                                        'cons_norm_behavior': bap_res[4],
                                        'perf_behavior': bap_res[2],
                                        'perf_norm_behavior': bap_res[7],
                                        'understaffing_behavior': bap_res[1],
                                        'understaffing_norm_behavior': bap_res[6],
                                        'undercover_naive': naive_expost_metrics['undercover_naive'],
                                        'undercover_norm_naive': naive_expost_metrics['undercover_norm_naive'],
                                        'cons_naive': naive_expost_metrics['cons_naive'],
                                        'cons_norm_naive': naive_expost_metrics['cons_norm_naive'],
                                        'perf_naive': naive_expost_metrics['perf_naive'],
                                        'perf_norm_naive': naive_expost_metrics['perf_norm_naive'],
                                        'understaffing_naive': naive_expost_metrics['understaffing_naive'],
                                        'understaffing_norm_naive': naive_expost_metrics['understaffing_norm_naive'],
                                        'shift_undercover_naive': str(naive_expost_metrics['shift_undercover_naive']),
                                        'shift_undercover_behavior': str([0.0 if abs(round(x, 3)) == 0 else round(x, 2) for x in bap_res[21]]),
                                        'perf_list_behavior': str(bap_res[18]),
                                        'perf_list_naive': str(naive_expost_metrics['perf_list_naive']),
                                        'cons_list_behavior': str(bap_res[17]),
                                        'cons_list_naive': str(naive_expost_metrics['cons_list_naive']),
                                        'p_list_behavior': str(bap_res[16]),
                                        'p_list_naive': str(naive_expost_metrics['p_list_naive']),
                                        'x_list_behavior': str(bap_res[19]),
                                        'x_list_naive': str(naive_expost_metrics['x_list_naive']),
                                        'r_list_behavior': str(bap_res[20]),
                                        'r_list_naive': str(naive_expost_metrics['r_list_naive']),
                                        'daily_undercover_behavior': str(daily_undercover_behavior),
                                        'daily_undercover_naive': str(naive_expost_metrics['daily_undercover_naive']),
                                        'results_ineq_sc_behavior': str(bap_res[22]),
                                        'results_ineq_sc_naive': str(naive_expost_metrics['results_ineq_sc_naive']),
                                        'spread_sc_behavior': bap_res[23],
                                        'spread_sc_naive': naive_expost_metrics['spread_sc_naive'],
                                        'load_share_sc_behavior': bap_res[24],
                                        'load_share_sc_naive': naive_expost_metrics['load_share_sc_naive'],
                                        'gini_sc_behavior': bap_res[25],
                                        'gini_sc_naive': naive_expost_metrics['gini_sc_naive'],
                                        'results_ineq_perf_behavior': str(bap_res[26]),
                                        'results_ineq_perf_naive': str(naive_expost_metrics['results_ineq_perf_naive']),
                                        'spread_perf_behavior': bap_res[27],
                                        'spread_perf_naive': naive_expost_metrics['spread_perf_naive'],
                                        'load_share_perf_behavior': bap_res[28],
                                        'load_share_perf_naive': naive_expost_metrics['load_share_perf_naive'],
                                        'gini_perf_behavior': bap_res[29],
                                        'gini_perf_naive': naive_expost_metrics['gini_perf_naive'],
                                        'shift_blocks_behavior': str(bap_res[30]),
                                        'shift_blocks_naive': str(naive_expost_metrics['shift_blocks_naive'])
                                    }
                                    
                                    row_data = [row_dict[h] for h in headers]
                                    writer.writerow(row_data)
                                    f.flush()
                                    
                                    ws.append(row_data)
                                    wb.save(excel_file)

if __name__ == "__main__":
    run_experiment()
